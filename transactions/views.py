import openpyxl
from django.db.models import F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, Side
from rest_framework import permissions, status, viewsets

from api.mixins import CustomPaginationMixin
from api.pagination import CustomPagination
from api.utils import api_response
from authentication.permissions import IsAdmin, IsCashier
from store.models import Store
from suppliers.models.supplier import Supplier
from transactions.models.transaction import Transaction
from transactions.models.transaction_item import TransactionItem
from transactions.serializers.transaction import TransactionSerializer, TransactionUpdateSerializer
from transactions.serializers.transaction_item import SupplierTransactionItemSerializer


class TransactionViewSet(CustomPaginationMixin, viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer
    pagination_class = CustomPagination
    permission_classes = [permissions.IsAuthenticated, (IsAdmin | IsCashier)]

    def get_queryset(self):
        queryset = Transaction.objects.all().order_by('-updated_at')
        user = self.request.user

        if getattr(user, 'role', None) == 'admin':
            queryset = queryset.filter(
                Q(is_saved=False) |
                Q(cashier_book_transactions__cashier_book__cashier=user)
            ).distinct()
        elif getattr(user, 'role', None) == 'cashier':
            queryset = queryset.filter(cashier_book_transactions__cashier_book__cashier=user)
        
        if self.action == 'list':
            search = self.request.query_params.get('search')
            cashier_id = self.request.query_params.get('cashier_id')
            start_date = self.request.query_params.get('start_date')
            end_date = self.request.query_params.get('end_date')
            transaction_status = self.request.query_params.get('transaction_status')
            payment = self.request.query_params.get('payment')

            if search:
                queryset = queryset.filter(
                    Q(code__icontains=search)
                )

            if cashier_id:
                queryset = queryset.filter(cashier_book_transactions__cashier_book__cashier_id=cashier_id)

            if start_date:
                queryset = queryset.filter(created_at__date__gte=start_date)

            if end_date:
                queryset = queryset.filter(created_at__date__lte=end_date)

            if transaction_status:
                status_list = [s.strip().lower() for s in transaction_status.split(',')]
                status_filter = []
                if 'saved' in status_list:
                    status_filter.append(True)
                if 'paid' in status_list:
                    status_filter.append(False)
                if status_filter:
                    queryset = queryset.filter(is_saved__in=status_filter)

            if payment:
                payment_list = [p.strip().lower() for p in payment.split(',')]
                queryset = queryset.filter(payment__in=payment_list)

        return queryset

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return TransactionUpdateSerializer
        return TransactionSerializer

    def perform_create(self, serializer):
        pay = serializer.validated_data.get('pay')
        if pay and pay != 0:
            serializer.save(paid_time=timezone.now())
        else:
            serializer.save()

    def perform_update(self, serializer):
        pay = serializer.validated_data.get('pay')
        
        if pay and pay != 0:
            serializer.save(paid_time=timezone.now())
        else:
            serializer.save()

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data, message="Transactions retrieved successfully")

        serializer = self.get_serializer(queryset, many=True)
        return api_response(
            status=status.HTTP_200_OK,
            success=True,
            message="Transactions retrieved successfully",
            data=serializer.data
        )

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return api_response(
            status=status.HTTP_201_CREATED,
            success=True,
            message="Transaction created successfully",
            data=response.data
        )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return api_response(
            status=status.HTTP_200_OK,
            success=True,
            message="Transaction retrieved successfully",
            data=serializer.data
        )

    def update(self, request, *args, **kwargs):
        super().update(request, *args, **kwargs)
        instance = self.get_object()
        serializer = TransactionSerializer(instance)
        return api_response(
            status=status.HTTP_200_OK,
            success=True,
            message="Transaction updated successfully",
            data=serializer.data
        )

    def supplier_products(self, request, supplier_id=None):
        user = request.user
        if getattr(user, 'role', None) != 'admin':
             return api_response(status=status.HTTP_403_FORBIDDEN, success=False, message="You do not have permission to access this resource.")
        
        if not supplier_id:
            return api_response(status=status.HTTP_400_BAD_REQUEST, success=False, message="Supplier ID is required.")

        queryset = TransactionItem.objects.filter(product_sku__supplier_id=supplier_id).select_related('transaction', 'product_sku__product', 'product_sku__product__category').order_by('-transaction__created_at')

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(transaction__paid_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction__paid_time__date__lte=end_date)
            
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(product_sku__product__name__icontains=search) |
                Q(product_sku__sku__icontains=search) |
                Q(transaction__code__icontains=search)
            )
            
        categories = request.query_params.get('categories')
        if categories:
            category_list = [c.strip() for c in categories.split(',')]
            queryset = queryset.filter(product_sku__product__category__id__in=category_list)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = SupplierTransactionItemSerializer(page, many=True)
            return self.get_paginated_response(serializer.data, message="Supplier products transactions retrieved successfully")

        serializer = SupplierTransactionItemSerializer(queryset, many=True)
        return api_response(status=status.HTTP_200_OK, success=True, message="Supplier products transactions retrieved successfully", data=serializer.data)

    def export_supplier_products(self, request, supplier_id=None):
        user = request.user
        if getattr(user, 'role', None) != 'admin':
             return api_response(status=status.HTTP_403_FORBIDDEN, success=False, message="You do not have permission to access this resource.")
        
        if not supplier_id:
            return api_response(status=status.HTTP_400_BAD_REQUEST, success=False, message="Supplier ID is required.")
            
        try:
            supplier = Supplier.objects.get(id=supplier_id)
        except Supplier.DoesNotExist:
            return api_response(status=status.HTTP_404_NOT_FOUND, success=False, message="Supplier not found.")

        queryset = TransactionItem.objects.filter(product_sku__supplier_id=supplier_id)

        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        if start_date:
            queryset = queryset.filter(transaction__paid_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction__paid_time__date__lte=end_date)
            
        search = request.data.get('search')
        if search:
            queryset = queryset.filter(
                Q(product_sku__product__name__icontains=search) |
                Q(product_sku__sku__icontains=search) |
                Q(transaction__code__icontains=search)
            )
            
        categories = request.data.get('categories')
        if categories:
            if isinstance(categories, str):
                category_list = [c.strip() for c in categories.split(',')]
                queryset = queryset.filter(product_sku__product__category__id__in=category_list)
            elif isinstance(categories, list):
                queryset = queryset.filter(product_sku__product__category__id__in=categories)

        payment_option = request.data.get('payment_option')
        if payment_option:
            if isinstance(payment_option, str):
                payment_option_list = [p.strip() for p in payment_option.split(',')]
                queryset = queryset.filter(product_sku__payment_option__in=payment_option_list)
            elif isinstance(payment_option, list):
                queryset = queryset.filter(product_sku__payment_option__in=payment_option)

        aggregated_data = queryset.values(
            sku=F('product_sku__sku'),
            product_name=F('product_sku__product__name'),
            payment_option=F('product_sku__payment_option'),
            selling_price=F('unit_price'),
            discount=Coalesce('supplier_discount', Value(0.0))
        ).annotate(
            sold_amounts=Sum('amount')
        ).order_by('product_name', 'selling_price', 'discount')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="supplier_products_export.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier Products"
        
        # Styles
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12) # For UMS Store
        header_font = Font(bold=True)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
                        
        store = Store.objects.first()
        store_name = store.name
        store_address = store.address
        store_phone = f"Telp. {store.phone}"

        # Title Section
        ws.merge_cells('A1:I1')
        ws['A1'] = "SALES REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment

        ws.merge_cells('A2:I2')
        ws['A2'] = store_name
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_alignment

        ws.merge_cells('A3:I3')
        ws['A3'] = store_address
        ws['A3'].alignment = center_alignment

        ws.merge_cells('A4:I4')
        ws['A4'] = store_phone
        ws['A4'].alignment = center_alignment

        # Metadata
        # Row 6: Start Date
        ws['A6'] = "Start Date :"
        ws['A6'].font = bold_font
        ws['B6'] = start_date if start_date else "-"

        # Row 7: End Date
        ws['A7'] = "End Date :"
        ws['A7'].font = bold_font
        ws['B7'] = end_date if end_date else "-"

        # Row 8: Supplier Code
        ws['A8'] = "Supplier Code :"
        ws['A8'].font = bold_font
        ws['B8'] = supplier.code

        # Row 9: Supplier Name
        ws['A9'] = "Supplier Name :"
        ws['A9'].font = bold_font
        ws['B9'] = supplier.name

        # Headers
        headers = ['No', 'SKU', 'Product Name', 'Payment Option', 'Selling Price', 'Supplier Discount', 'Netto', 'Sold Amounts', 'Total Netto']
        header_row = 11
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header_title)
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # Data
        row_num = 12
        
        for idx, item in enumerate(aggregated_data, 1):
            selling_price = item['selling_price']
            discount = item['discount']
            sold_amounts = item['sold_amounts']
            
            # Write cells
            no_cell = ws.cell(row=row_num, column=1, value=idx)
            no_cell.border = border
            no_cell.alignment = center_alignment
            sku_cell = ws.cell(row=row_num, column=2, value=item['sku'])
            sku_cell.border = border
            sku_cell.alignment = center_alignment
            ws.cell(row=row_num, column=3, value=item['product_name']).border = border
            # Make Payment Option capitalized if simpler
            po_val = item['payment_option'].title() if item['payment_option'] else '-'
            payment_cell = ws.cell(row=row_num, column=4, value=po_val)
            payment_cell.border = border
            payment_cell.alignment = center_alignment
            
            selling_price_cell = ws.cell(row=row_num, column=5, value=selling_price)
            selling_price_cell.border = border
            selling_price_cell.alignment = center_alignment
            selling_price_cell.number_format = '"Rp" #,##0'
            
            # Use number for discount and format as percentage
            discount_val = discount / 100
            discount_cell = ws.cell(row=row_num, column=6, value=discount_val)
            discount_cell.border = border
            discount_cell.alignment = center_alignment
            discount_cell.number_format = '0%'

            # Netto Calculation Formula: Selling Price * (1 - Discount)
            # E{row_num} * (1 - F{row_num})
            netto_formula = f'=E{row_num}*(1-F{row_num})'
            netto_cell = ws.cell(row=row_num, column=7, value=netto_formula)
            netto_cell.border = border
            netto_cell.alignment = center_alignment
            netto_cell.number_format = '"Rp" #,##0'
            
            sold_cell = ws.cell(row=row_num, column=8, value=sold_amounts)
            sold_cell.border = border
            sold_cell.alignment = center_alignment
            
            # Total Netto Calculation Formula: Netto * Sold Amounts
            # G{row_num} * H{row_num}
            total_netto_formula = f'=G{row_num}*H{row_num}'
            total_netto_cell = ws.cell(row=row_num, column=9, value=total_netto_formula)
            total_netto_cell.border = border
            total_netto_cell.alignment = center_alignment
            total_netto_cell.number_format = '"Rp" #,##0'
            
            row_num += 1

        # Total Row
        # Merge A to G (1 to 7)
        ws.merge_cells(f'A{row_num}:G{row_num}')
        total_label_cell = ws.cell(row=row_num, column=1, value="Total")
        total_label_cell.font = bold_font
        total_label_cell.alignment = center_alignment
        
        # Apply borders to the footer row
        rows = ws[f'A{row_num}:I{row_num}']
        for cell in rows[0]:
            cell.border = border
            
        # Set values
        # Sold Total Formula: SUM(H12:H{last_row})
        sold_total_formula = f'=SUM(H12:H{row_num-1})'
        sold_total_cell = ws.cell(row=row_num, column=8, value=sold_total_formula)
        sold_total_cell.font = bold_font
        sold_total_cell.alignment = center_alignment
        sold_total_cell.border = border
        
        # Netto Total Formula: SUM(I12:I{last_row})
        netto_total_formula = f'=SUM(I12:I{row_num-1})'
        netto_total_cell = ws.cell(row=row_num, column=9, value=netto_total_formula)
        netto_total_cell.font = bold_font
        netto_total_cell.alignment = center_alignment
        netto_total_cell.border = border
        netto_total_cell.number_format = '"Rp" #,##0'

        # Adjust column widths
        ws.column_dimensions['A'].width = 15

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15 # Payment Option
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15

        wb.save(response)
        return response

    def export_supplier_sales_report(self, request):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Supplier_Sales_Report_{timezone.now().strftime("%Y%m%d%H%M%S")}.xlsx"'

        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        queryset = TransactionItem.objects.filter(transaction__paid_time__isnull=False)

        if start_date:
            queryset = queryset.filter(transaction__paid_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction__paid_time__date__lte=end_date)

        aggregated_data = queryset.values(
            'product_sku__supplier__code', 'product_sku__supplier__name'
        ).annotate(
            supplier_code=F('product_sku__supplier__code'),
            supplier_name=F('product_sku__supplier__name'),
            product_sold=Sum('amount'),
            sales_total=Sum(F('unit_price') * F('amount'))
        ).order_by('product_sku__supplier__name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Styles
        title_font = Font(name='Calibri', size=14, bold=True)
        subtitle_font = Font(name='Calibri', size=12, bold=True)
        bold_font = Font(name='Calibri', bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        store = Store.objects.first()
        store_name = store.name if store else "UMS Store"
        store_address = store.address if store else "Sukoharjo"
        store_phone = f"Telp. {store.phone}" if store else "Telp. -"

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "SALES REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment

        ws.merge_cells('A2:E2')
        ws['A2'] = store_name
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_alignment

        ws.merge_cells('A3:E3')
        ws['A3'] = store_address
        ws['A3'].alignment = center_alignment

        ws.merge_cells('A4:E4')
        ws['A4'] = store_phone
        ws['A4'].alignment = center_alignment

        # Dates
        ws['A6'] = "Start Date :"
        ws['A6'].font = bold_font
        ws['B6'] = start_date if start_date else "-"

        ws['A7'] = "End Date :"
        ws['A7'].font = bold_font
        ws['B7'] = end_date if end_date else "-"

        # Headers
        headers = ['No', 'Supplier Code', 'Supplier Name', 'Product Sold', 'Sales Total']
        header_row = 9
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header_title)
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_alignment

        # Data
        row_num = 10
        total_product_sold = 0
        total_sales_total = 0

        for idx, item in enumerate(aggregated_data, 1):
            product_sold = item['product_sold'] or 0
            sales_total = item['sales_total'] or 0
            
            total_product_sold += product_sold
            total_sales_total += sales_total

            ws.cell(row=row_num, column=1, value=idx).border = border
            ws.cell(row=row_num, column=1).alignment = center_alignment

            ws.cell(row=row_num, column=2, value=item['supplier_code'] or '-').border = border
            ws.cell(row=row_num, column=2).alignment = center_alignment

            ws.cell(row=row_num, column=3, value=item['supplier_name'] or '-').border = border
            
            ws.cell(row=row_num, column=4, value=product_sold).border = border
            ws.cell(row=row_num, column=4).alignment = center_alignment
            
            sales_cell = ws.cell(row=row_num, column=5, value=sales_total)
            sales_cell.border = border
            sales_cell.alignment = center_alignment
            sales_cell.number_format = '"Rp" #,##0'

            row_num += 1

        # Footer
        ws.merge_cells(f'A{row_num}:C{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value="Total")
        total_cell.font = bold_font
        total_cell.alignment = center_alignment
        
        # Apply border for merged cells
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = border

        sold_cell = ws.cell(row=row_num, column=4, value=total_product_sold)
        sold_cell.font = bold_font
        sold_cell.alignment = center_alignment
        sold_cell.border = border

        sales_cell = ws.cell(row=row_num, column=5, value=total_sales_total)
        sales_cell.font = bold_font
        sales_cell.alignment = center_alignment
        sales_cell.border = border
        sales_cell.number_format = '"Rp" #,##0'

        # Widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20

        wb.save(response)
        return response

    def export_product_category_sales(self, request):
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        queryset = TransactionItem.objects.all()

        if start_date:
            queryset = queryset.filter(transaction__paid_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction__paid_time__date__lte=end_date)

        aggregated_data = queryset.annotate(
            category_name=Coalesce('product_sku__product__category__name', Value('No Category'))
        ).values(
            'category_name'
        ).annotate(
            product_sold=Sum('amount'),
            sales_total=Sum(F('unit_price') * F('amount'))
        ).order_by('category_name')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="product_category_sales_report.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Category Sales"
        
        # Styles
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
                        
        store = Store.objects.first()
        store_name = store.name if store else "UMS Store"
        store_address = store.address if store else "-"
        store_phone = f"Telp. {store.phone}" if store else "-"

        # Title Section
        ws.merge_cells('A1:D1')
        ws['A1'] = "SALES REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment

        ws.merge_cells('A2:D2')
        ws['A2'] = store_name
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_alignment

        ws.merge_cells('A3:D3')
        ws['A3'] = store_address
        ws['A3'].alignment = center_alignment

        ws.merge_cells('A4:D4')
        ws['A4'] = store_phone
        ws['A4'].alignment = center_alignment

        # Metadata
        ws['A6'] = "Start Date :"
        ws['A6'].font = bold_font
        ws['B6'] = start_date if start_date else "-"

        ws['A7'] = "End Date :"
        ws['A7'].font = bold_font
        ws['B7'] = end_date if end_date else "-"

        # Headers
        headers = ['No', 'Category Name', 'Product Sold', 'Sales Total']
        header_row = 9
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header_title)
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_alignment

        # Data
        row_num = 10
        total_product_sold = 0
        total_sales_total = 0

        for idx, item in enumerate(aggregated_data, 1):
            product_sold = item['product_sold'] or 0
            sales_total = item['sales_total'] or 0
            
            total_product_sold += product_sold
            total_sales_total += sales_total

            ws.cell(row=row_num, column=1, value=idx).border = border
            ws.cell(row=row_num, column=1).alignment = center_alignment

            ws.cell(row=row_num, column=2, value=item['category_name']).border = border
            ws.cell(row=row_num, column=2).alignment = center_alignment
            
            ws.cell(row=row_num, column=3, value=product_sold).border = border
            ws.cell(row=row_num, column=3).alignment = center_alignment
            
            sales_cell = ws.cell(row=row_num, column=4, value=sales_total)
            sales_cell.border = border
            sales_cell.alignment = center_alignment
            sales_cell.number_format = '"Rp" #,##0'

            row_num += 1

        # Footer
        ws.merge_cells(f'A{row_num}:B{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value="Total")
        total_cell.font = bold_font
        total_cell.alignment = center_alignment
        
        # Apply border for merged cells
        for col in range(1, 3):
            ws.cell(row=row_num, column=col).border = border

        sold_cell = ws.cell(row=row_num, column=3, value=total_product_sold)
        sold_cell.font = bold_font
        sold_cell.alignment = center_alignment
        sold_cell.border = border

        sales_cell = ws.cell(row=row_num, column=4, value=total_sales_total)
        sales_cell.font = bold_font
        sales_cell.alignment = center_alignment
        sales_cell.border = border
        sales_cell.number_format = '"Rp" #,##0'

        # Widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

        wb.save(response)
        return response
